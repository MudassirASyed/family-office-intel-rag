"""
Exports data/records.json to a client-facing CSV (and XLSX) - the
"structured dataset file" deliverable the brief requires as an actual
attached file, not a link.

Column order is deliberate: actionable fields first (who to contact,
how, what they invest in), verification/audit fields after (basis for
trust, not the first thing a fund manager needs). sources_checked and
evidence (dicts, per source class) are flattened into readable strings
here rather than shipped as raw JSON - the brief is explicit that a raw
data dump is not a customer-facing artifact, and that applies to this
file too, not just the RAG UI.

Usage:
    python scripts/export_dataset.py
Writes data/family_office_dataset.csv and .xlsx
"""
import sys
import os
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import csv
import json

RECORDS_PATH = "data/records.json"
CSV_OUT = "data/family_office_dataset.csv"
XLSX_OUT = "data/family_office_dataset.xlsx"

COLUMNS = [
    "name", "firm_type",
    "principal_name", "principal_title",
    "principal_email", "principal_email_verified", "principal_phone",
    "principal_linkedin", "corporate_linkedin", "website",
    "city", "state", "country",
    "aum", "investment_thesis", "investing_sectors", "description",
    "recent_activity", "recent_activity_date",
    "firm_qualifies", "qualification_evidence",
    "confidence", "verification_sources", "verification_evidence",
    "date_verified", "notes",
]

FIRM_TYPE_LABELS = {
    "single_family_office": "Single-Family Office",
    "multi_family_office": "Multi-Family Office",
    "unclear": "Unclear",
}


def flatten(record: dict) -> dict:
    sources_checked = record.get("sources_checked", {}) or {}
    evidence = record.get("evidence", {}) or {}

    verified_source_classes = [k for k, v in sources_checked.items() if v]
    evidence_str = " | ".join(f"{k}: {v}" for k, v in evidence.items())

    row = dict(record)
    row["firm_type"] = FIRM_TYPE_LABELS.get(record.get("firm_type"), record.get("firm_type", ""))
    row["verification_sources"] = ", ".join(verified_source_classes)
    row["verification_evidence"] = evidence_str
    return {col: row.get(col, "") for col in COLUMNS}


def main():
    with open(RECORDS_PATH, encoding="utf-8") as f:
        records = json.load(f)

    rows = [flatten(r) for r in records]
    rows.sort(key=lambda r: r["name"])

    with open(CSV_OUT, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=COLUMNS)
        writer.writeheader()
        writer.writerows(rows)
    print(f"Wrote {len(rows)} records to {CSV_OUT}")

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = "Family Offices"
        ws.append(COLUMNS)
        for cell in ws[1]:
            cell.font = Font(bold=True)
        for row in rows:
            ws.append([row[col] for col in COLUMNS])
        ws.freeze_panes = "A2"
        for i, col in enumerate(COLUMNS, start=1):
            width = min(max(len(col), 14), 40)
            ws.column_dimensions[get_column_letter(i)].width = width
        wb.save(XLSX_OUT)
        print(f"Wrote {len(rows)} records to {XLSX_OUT}")
    except ImportError:
        print("openpyxl not installed - skipped XLSX export (CSV is still valid on its own).")


if __name__ == "__main__":
    main()
